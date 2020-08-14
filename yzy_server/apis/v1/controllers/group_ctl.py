import logging
import openpyxl
import ipaddress
import datetime
from yzy_server.database import apis as db_api
from common.utils import create_uuid, create_md5
from common.errcode import get_error_result
from common import constants
from common.config import SERVER_CONF
from common.utils import terminal_post


logger = logging.getLogger(__name__)


class GroupController(object):

    def _check_params(self, data):
        if not data:
            return False
        name = data.get('name', '')
        group_type = data.get('group_type', '')
        if not (name and group_type):
            return False
        logger.info("check params ok")
        return True

    def change_group_uuid(self, terminals, start_ip, end_ip, group_uuid):
        if terminals:
            for terminal in terminals:
                flag_a = ipaddress.ip_network(terminal.ip).compare_networks(ipaddress.ip_network(start_ip))
                flag_b = ipaddress.ip_network(terminal.ip).compare_networks(ipaddress.ip_network(end_ip))
                if flag_a != -1 and flag_b != 1:
                    terminal.group_uuid = group_uuid
                    terminal.soft_update()

    def create_group(self, data):
        """
        创建分组，包括教学分组和用户分组
        """
        if not self._check_params(data):
            return get_error_result("ParamError")

        if constants.EDUCATION_DESKTOP == data.get('group_type', constants.EDUCATION_DESKTOP):
            subnet = db_api.get_subnet_by_uuid(data['subnet_uuid'])
            if not subnet:
                logger.error("subnet: %s not exist", data['subnet_uuid'])
                return get_error_result("SubnetNotExist")

        groups = db_api.get_group_with_all({'name': data['name'], 'group_type': data['group_type']})
        if groups:
            return get_error_result("GroupAlreadyExists", name=data['name'])
        # add education group
        group_uuid = create_uuid()
        group_value = {
            "uuid": group_uuid,
            "group_type": data['group_type'],
            "name": data['name'],
            "desc": data['desc'],
            "network_uuid": data.get('network_uuid', ''),
            "subnet_uuid": data.get('subnet_uuid', ''),
            "start_ip": data.get('start_ip', ''),
            "end_ip": data.get('end_ip', '')
        }
        try:
            db_api.create_group(group_value)
            logger.info("create group %s success", data['name'])
        except Exception as e:
            logging.info("insert group info to db failed:%s", e)
            return get_error_result("GroupCreateError", name=data['name'])
        if data['group_type'] == constants.EDUCATION_DESKTOP:
            terminals = db_api.get_terminal_with_all({'group_uuid': None})
            self.change_group_uuid(terminals, data.get('start_ip', ''), data.get('end_ip', ''), group_uuid)
        return get_error_result("Success", group_value)

    def delete_group(self, group_uuid):
        group = db_api.get_group_with_first({"uuid": group_uuid})
        if not group:
            logger.error("group: %s not exist" % group_uuid)
            return get_error_result("GroupNotExists", name='')
        # 教学分组如果在使用中，不能删除
        if constants.EDUCATION_DESKTOP == group.group_type:
            desktop = db_api.get_desktop_with_all({"group_uuid": group_uuid})
            if desktop:
                logger.error("group already in use", group_uuid)
                return get_error_result("GroupInUse", name=group.name)
        # 个人分组删除时需要删除分组中的用户
        if constants.PERSONAL_DEKSTOP == group.group_type:
            users = db_api.get_group_user_with_all({'group_uuid': group.uuid})
            for user in users:
                binds = db_api.get_instance_with_all({"user_uuid": user.uuid})
                for bind in binds:
                    bind.user_uuid = ''
                    bind.soft_update()
                user.soft_delete()
            desktops = db_api.get_personal_desktop_with_all({"group_uuid": group_uuid})
            for desktop in desktops:
                desktop.group_uuid = ""
                desktop.soft_update()
        group.soft_delete()
        logger.info("delete group %s success", group_uuid)
        ret = terminal_post("/api/v1/terminal/task", {"handler": "WebTerminalHandler",
                                                         "command": "delete_group",
                                                         "data": {
                                                                    "group_uuid": group_uuid
                                                                  }
                                                        }
                            )
        return ret

    def update_group(self, data):
        group_uuid = data.get('uuid', '')
        group = db_api.get_group_with_first({"uuid": group_uuid})
        if not group:
            logger.error("group: %s not exist" % group_uuid)
            return get_error_result("GroupNotExists", name='')
        if group.group_type == 1:
            desktop = db_api.get_desktop_with_all({'group_uuid': group_uuid})
            if desktop:
                if 'subnet_uuid' in data['value'] and data['value']['subnet_uuid'] != group.subnet_uuid:
                    return get_error_result("GroupSubnetError")
            start_ip = data.get('value').get('start_ip')
            end_ip = data.get('value').get('end_ip')
            if group.start_ip != start_ip or group.end_ip != end_ip:
                old_terminals = db_api.get_terminal_with_all({'group_uuid': group_uuid})
                for terminal in old_terminals:
                    terminal.group_uuid = None
                    terminal.updated_at = datetime.datetime.utcnow()
                    terminal.soft_update()
                new_terminals = db_api.get_terminal_with_all({'group_uuid': None})
                self.change_group_uuid(new_terminals, start_ip, end_ip, group_uuid)
        try:
            group.update(data['value'])
            group.soft_update()
        except Exception as e:
            logger.error("update group:%s failed:%s", group_uuid, e)
            return get_error_result("GroupUpdateError", name=group.name)
        logger.info("update group:%s success", group_uuid)
        return get_error_result("Success")


class GroupUserController(object):

    def _check_params(self, data):
        if not data:
            return False
        group_uuid = data.get('group_uuid', '')
        user_name = data.get('user_name', '')
        passwd = data.get('passwd', '')
        if not (group_uuid and user_name and passwd):
            return False
        logger.info("check params ok")
        return True

    def create_user(self, data):
        if not self._check_params(data):
            return get_error_result("ParamError")

        group = db_api.get_group_with_first({'uuid': data['group_uuid']})
        if not group:
            logger.error("group: %s not exist", data['group_uuid'])
            return get_error_result("GroupNotExists", name='')

        # add user to group
        user_value = {
            "uuid": create_uuid(),
            "group_uuid": data['group_uuid'],
            "user_name": data['user_name'],
            "passwd": create_md5(data['passwd']),
            "name": data.get('name', ''),
            "phone": data.get('phone', ''),
            "email": data.get('email', ''),
            "enabled": data.get('enabled', True)
        }
        try:
            db_api.create_group_user(user_value)
            logger.info("create group user %s success", data['user_name'])
        except Exception as e:
            logging.info("insert group user info to db failed:%s", e)
            return get_error_result("GroupUserCreateError", user_name=data['user_name'])
        return get_error_result("Success", user_value)

    def multi_create_user(self, data):
        """
        如果用户名冲突，则失败计数加1
        """
        group = db_api.get_group_with_first({'uuid': data['group_uuid']})
        if not group:
            logger.error("group: %s not exist", data['group_uuid'])
            return get_error_result("GroupNotExists", name='')
        success_num = 0
        failed_num = 0
        postfix = data['postfix']
        postfix_start = data['postfix_start']
        for i in range(data['user_num']):
            # 桌面名称是前缀加几位数字
            if len(str(postfix_start)) < postfix:
                post = '0' * (postfix - len(str(postfix_start))) + str(postfix_start)
            else:
                post = str(postfix_start)
            user_name = data['prefix'] + post
            if db_api.get_group_user_with_first({'user_name': user_name}):
                logger.info("multi create user failed, the user %s already exists", user_name)
                failed_num += 1
                postfix_start += 1
                continue
            user_uuid = create_uuid()
            user_value = {
                "uuid": user_uuid,
                "group_uuid": data['group_uuid'],
                "user_name": user_name,
                "passwd": create_md5(data['passwd']),
                "name": data.get('name', ''),
                "phone": data.get('phone', ''),
                "email": data.get('email', ''),
                "enabled": data.get('enabled', True)
            }
            postfix_start += 1
            try:
                db_api.create_group_user(user_value)
                logger.info("create user:%s success", user_name)
            except Exception as e:
                logging.info("insert group user info to db failed:%s", e)
                failed_num += 1
                postfix_start += 1
                continue
            success_num += 1
        logger.info("multi create group user success")
        return get_error_result("Success", {"failed_num": failed_num, "success_num": success_num})

    def delete_user(self, user_uuid):
        user = db_api.get_group_user_with_first({'uuid': user_uuid})
        if not user:
            logger.error("user: %s not exist" % user_uuid)
            return get_error_result("GroupUserNotExists", name='')
        binds = db_api.get_instance_with_all({"user_uuid": user.uuid})
        for bind in binds:
            bind.user_uuid = ''
            bind.soft_update()
        user.soft_delete()
        logger.info("delete group user %s success", user_uuid)
        return get_error_result("Success")

    def enable_user(self, user_uuid):
        user = db_api.get_group_user_with_first({'uuid': user_uuid})
        if not user:
            logger.error("user %s not exist", user_uuid)
            return get_error_result("GroupUserNotExists", name='')
        try:
            user.enabled = True
            user.soft_update()
        except Exception as e:
            logger.error("enable user %s failed:%s", user.user_name, e)
            return get_error_result("GroupUserDisableError", user_name=user.user_name)
        logger.info("enable user %s success", user.user_name)
        return get_error_result("Success")

    def disable_user(self, user_uuid):
        user = db_api.get_group_user_with_first({'uuid': user_uuid})
        if not user:
            logger.error("user %s not exist", user_uuid)
            return get_error_result("GroupUserNotExists", name='')
        try:
            user.enabled = False
            user.soft_update()
        except Exception as e:
            logger.error("disable user %s failed:%s", user.user_name, e)
            return get_error_result("GroupUserDisableError", user_name=user.user_name)
        logger.info("disable user %s success", user.user_name)
        return get_error_result("Success")

    def reset_passwd(self, user_uuid):
        user = db_api.get_group_user_with_first({'uuid': user_uuid})
        if not user:
            logger.error("user %s not exist", user_uuid)
            return get_error_result("GroupUserNotExists", name='')
        try:
            user.passwd = create_md5("123456")
            user.soft_update()
        except Exception as e:
            logger.error("reset user %s passwd failed:%s", user.user_name, e)
            return get_error_result("GroupUserUpdateError", user_name=user.user_name)
        logger.info("reset user %s passwd success", user.user_name)
        return get_error_result("Success")

    def move_user(self, data):
        group = db_api.get_group_with_first({"uuid": data['group_uuid']})
        if not group:
            logger.error("group: %s not exist" % data['group_uuid'])
            return get_error_result("GroupNotExists", name='')
        for user in data['users']:
            user = db_api.get_group_user_with_first({'uuid': user['uuid']})
            if not user:
                logger.error("user %s not exist", user['user_name'])
                return get_error_result("GroupUserNotExists", name=user['user_name'])
            if user.group_uuid == group.uuid:
                logger.info("user %s already in group %s, skip", user.user_name, group.name)
                continue
            try:
                user.group_uuid = data['group_uuid']
                # 移动后，原先的跟桌面绑定关系要删除
                binds = db_api.get_instance_with_all({"user_uuid": user.uuid})
                for bind in binds:
                    bind.user_uuid = ''
                    bind.soft_update()
                user.soft_update()
            except Exception as e:
                logger.error("move user %s to group %s failed:%s", user.user_name, group.name, e)
                return get_error_result("GroupUserMoveError", user_name=user.user_name, group=group.name)
            logger.info("move user %s to group %s success", user.user_name, group.name)
        return get_error_result("Success")

    def update_user(self, data):
        user_uuid = data.get('uuid', '')
        user = db_api.get_group_user_with_first({"uuid": user_uuid})
        if not user:
            logger.error("user: %s not exist" % user_uuid)
            return get_error_result("GroupUserNotExists", name='')
        if data['value'].get('passwd'):
            data['value']['passwd'] = create_md5(data['value']['passwd'])
        # if not data['value'].get('name'):
        #     data['value'].pop('name')
        # if not data['value'].get('email'):
        #     data['value'].pop('email')
        # if not data['value'].get('phone'):
        #     data['value'].pop('phone')
        try:
            user.update(data['value'])
            user.soft_update()
        except Exception as e:
            logger.error("update group user:%s failed:%s", user_uuid, e)
            return get_error_result("GroupUserUpdateError", user_name=user.user_name)
        logger.info("update group user:%s success", user_uuid)
        return get_error_result("Success")

    def export_user(self, data):
        """
        :param data:
            {
                "filename": "file1",
                "users": [
                        {
                            "uuid": "",
                            "user_name": ""
                        }
                    ]
            }
        :return:
        """
        logger.info("begin to export user:%s", data)
        users = data.get('users', [])
        filename = '%s.xlsx' % data['filename']
        filepath = '/root/%s' % filename
        book = openpyxl.Workbook()
        ws = book.active
        ws.title = 'users'
        lines = list()
        info = ["用户名", "状态", "所属分组", "姓名", "邮箱", "电话"]
        lines.append(info)
        for user in users:
            user = db_api.get_group_user_with_first({'uuid': user['uuid']})
            if not user:
                continue
                # return build_result("GroupUserNotExists", name=user['user_name'])
            if user.enabled:
                status = "启用"
            else:
                status = "禁用"
            user_info = [user.user_name, status, user.group.name, user.email, user.phone]
            lines.append(user_info)
        try:
            for line in lines:
                ws.append(line)
            book.save(filepath)
        except Exception as e:
            logger.error("write xlsx file failed:%s", e)
            return get_error_result("GroupUserExportError", file=filename)
        node = db_api.get_controller_node()
        bind = SERVER_CONF.addresses.get_by_default('server_bind', '')
        if bind:
            port = bind.split(':')[-1]
        else:
            port = constants.SERVER_DEFAULT_PORT
        endpoint = 'http://%s:%s' % (node.ip, port)
        url = '%s/api/v1/group/user/download?path=%s' % (endpoint, filepath)
        logger.info("export user to file %s success", filename)
        return get_error_result("Success", {"url": url})

    def import_user(self, data):
        file = data['filepath']
        enabled = data['enabled']
        logger.info("begin to import user, file:%s", file)
        success_num = 0
        failed_num = 0
        book = openpyxl.load_workbook(file)
        ws = book.active
        for row in ws.iter_rows(min_row=2, max_col=6):
            flag = True
            values = dict()
            for cell in row:
                if 1 == cell.column:
                    user_name = cell.value
                    if not user_name:
                        failed_num += 1
                        flag = False
                        logger.error("the user_name can not be blank, skip", user_name)
                        break
                    user = db_api.get_group_user_with_first({"user_name": user_name})
                    if user:
                        failed_num += 1
                        flag = False
                        logger.error("the user %s already exist in db, skip", user_name)
                        break
                    values['user_name'] = user_name
                elif 2 == cell.column:
                    passwd = cell.value
                    if not passwd:
                        failed_num += 1
                        flag = False
                        logger.error("the password can not be blank")
                        break
                    if len(str(passwd)) < 6 or len(str(passwd)) > 16:
                        failed_num += 1
                        flag = False
                        logger.error("the password len must between 6~16")
                        break
                    values['passwd'] = create_md5(cell.value)
                elif 3 == cell.column:
                    group_name = cell.value
                    group = db_api.get_group_with_first({"name": group_name})
                    if not group:
                        failed_num += 1
                        flag = False
                        logger.error("the group %s not exist, skip", group_name)
                        break
                    else:
                        values['group_uuid'] = group.uuid
                elif 4 == cell.column:
                    name = cell.value
                    if name:
                        values['name'] = str(name)
                elif 5 == cell.column:
                    email = cell.value
                    if email:
                        values['email'] = str(email)
                elif 6 == cell.column:
                    phone = cell.value
                    if phone:
                        values['phone'] = str(phone)
                else:
                    break
            if flag:
                values['enabled'] = int(enabled)
                values['uuid'] = create_uuid()
                db_api.create_group_user(values)
                success_num += 1
        return get_error_result("Success", {"failed_num": failed_num, "success_num": success_num})
